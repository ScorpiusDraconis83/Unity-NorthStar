# Copyright (c) Meta Platforms, Inc. and affiliates.
import openpyxl as xl
import xml.etree.ElementTree as ET
"""
Converts Xl doc to xml file for dialouge system
Run this from your assets folder
"""
path = "NorthStar/Scripts/Utils/NorthStar Localisation Sheet.xlsx"
wb = xl.load_workbook(path)
sheet = wb.active


dialogueCollums = {}
cell = sheet.cell(1,1)
i = 1
while cell.value != None:
    print(cell.value)
    if "Dialogue" in cell.value:
        dialogueCollums[i] = cell.value.replace(" Dialogue", "")
    i+=1
    cell = sheet.cell(1,i)
print(dialogueCollums)


def ParseRow(row):
    element = ET.Element("TextObject")
    element.attrib["Id"] = sheet.cell(row,1).value
    for i in dialogueCollums.keys():
        textElement = ET.Element("Text")
        textElement.attrib["Language"] = dialogueCollums[i]
        textElement.text = sheet.cell(row,i).value
        element.append(textElement)
    return element



root = ET.Element("TextData")

langElement = ET.Element("SupportedLanguages")
for lang in dialogueCollums.values():
    element = ET.Element("Language")
    element.attrib["Value"] = lang
    langElement.append(element)
root.append(langElement)

rowId = 1
cell = sheet.cell(rowId,1)
while cell.value != None:
    if "BEAT" not in cell.value and "Line ID" not in cell.value:
        root.append(ParseRow(rowId))
    rowId += 1
    cell = sheet.cell(rowId,1)
ET.indent(root, space="\t", level=0)
tree = ET.ElementTree(root)


with open("NorthStar/Scripts/Utils/TextData.xml","wb") as f:
    tree.write(f,method="xml",encoding="UTF-8")
